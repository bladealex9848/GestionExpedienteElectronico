import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from metadata_extractor import get_file_metadata, get_pdf_pages
import xlwings as xw
import re

def extract_metadata_from_existing_index(folder_path):
    for file in os.listdir(folder_path):
        if 'indice' in file.lower() and file.endswith(('.xlsx', '.xlsm', '.xls')):
            index_path = os.path.join(folder_path, file)
            wb = load_workbook(index_path, read_only=True)
            ws = wb.active
            
            metadata = {
                'Ciudad': ws['B2'].value,  # Cambiado de B3 a B2
                'Despacho Judicial': ws['B3'].value,
                'Serie o Subserie documental': ws['B4'].value,
                'No. Radicación del Proceso': ws['B5'].value,
                'Partes Procesales (Parte A)': ws['B6'].value,
                'Partes Procesales (Parte B)': ws['B7'].value,
                'Terceros Intervinientes': ws['B8'].value,
                'Cuaderno': ws['B9'].value,
                'Expediente Físico': ws['J3'].value,
                'No. Carpetas': ws['J5'].value,
                'No. Carpetas Digitalizadas': ws['J6'].value
            }
            
            # Extraer las fechas de creación de documento del índice anterior
            existing_dates = {}
            for row in ws.iter_rows(min_row=12, values_only=True):
                if row[0] and row[1]:  # Nombre del documento y fecha de creación
                    existing_dates[row[0]] = row[1]
            
            metadata['existing_dates'] = existing_dates
            
            return metadata
    
    return None

def generate_index_from_scratch(folder_path, existing_metadata=None):
    files = [f for f in os.listdir(folder_path) 
             if os.path.isfile(os.path.join(folder_path, f))
             and not f.startswith('.') 
             and os.path.splitext(f)[1] != '' 
             and not ('indice' in f.lower() and f.endswith(('.xlsx', '.xlsm', '.xls')))]
    
    files.sort(key=lambda x: os.path.getmtime(os.path.join(folder_path, x)))

    data = []
    current_page = 1
    existing_dates = existing_metadata.get('existing_dates', {}) if existing_metadata else {}

    for i, filename in enumerate(files, start=1):
        file_path = os.path.join(folder_path, filename)
        metadata = get_file_metadata(file_path)
        
        num_pages = get_pdf_pages(file_path) if metadata['extension'].lower() == '.pdf' else 1
        
        doc_name = re.sub(r'^\d+', '', os.path.splitext(filename)[0])
        
        # Usar la fecha del índice anterior si está disponible, de lo contrario usar la fecha de modificación
        creation_date = existing_dates.get(doc_name, metadata['modification_date'])
        
        data.append({
            'Nombre Documento': doc_name,
            'Fecha Creación Documento': creation_date,
            'Fecha Incorporación Expediente': datetime.now().strftime('%Y-%m-%d'),
            'Orden Documento': i,
            'Número Páginas': num_pages,
            'Página Inicio': current_page,
            'Página Fin': current_page + num_pages - 1,
            'Formato': metadata['extension'][1:],
            'Tamaño': metadata['size'],
            'Origen': 'Digitalizado' if metadata['extension'].lower() == '.pdf' else 'Electrónico',
            'Observaciones': ''
        })
        current_page += num_pages

    df = pd.DataFrame(data)
    return df

def generate_index_from_template(folder_path, template_path):
    # Extraer metadatos del índice existente si lo hay
    existing_metadata = extract_metadata_from_existing_index(folder_path)

    # Generar el índice
    df = generate_index_from_scratch(folder_path, existing_metadata)

    try:
        # Usar xlwings para manejar el archivo con macros
        with xw.App(visible=False) as app:
            wb = app.books.open(template_path)
            ws = wb.sheets[0]
            
            # Llenar la plantilla con los metadatos existentes
            if existing_metadata:
                ws.range('B2').value = existing_metadata.get('Ciudad', '')  # Cambiado a B2
                ws.range('B3').value = existing_metadata.get('Despacho Judicial', '')
                ws.range('B4').value = existing_metadata.get('Serie o Subserie documental', '')
                ws.range('B5').value = existing_metadata.get('No. Radicación del Proceso', '')
                ws.range('B6').value = existing_metadata.get('Partes Procesales (Parte A)', '')
                ws.range('B7').value = existing_metadata.get('Partes Procesales (Parte B)', '')
                ws.range('B8').value = existing_metadata.get('Terceros Intervinientes', '')
                ws.range('B9').value = existing_metadata.get('Cuaderno', '')
                ws.range('J3').value = existing_metadata.get('Expediente Físico', '')
                ws.range('J5').value = existing_metadata.get('No. Carpetas', '')
                ws.range('J6').value = existing_metadata.get('No. Carpetas Digitalizadas', '')

            # Llenar los datos del índice
            ws.range('A12').options(index=False, headers=False).value = df.values
            
            # Guardar y cerrar
            output_path = os.path.join(folder_path, "000IndiceElectronicoC01.xlsm")
            wb.save(output_path)
            wb.close()
        
        return output_path

    except Exception as e:
        print(f"Error al usar xlwings: {str(e)}")
        print("Intentando con openpyxl (sin macros)...")      

        # Si xlwings falla, usar openpyxl como alternativa (sin macros)
        wb = load_workbook(template_path, keep_vba=True)
        ws = wb.active

        # Llenar la plantilla con los metadatos existentes
        if existing_metadata:
            metadata_fields = ["Ciudad", "Despacho Judicial", "Serie o Subserie documental", 
                               "No. Radicación del Proceso", "Partes Procesales (Parte A)", 
                               "Partes Procesales (Parte B)", "Terceros Intervinientes", "Cuaderno"]
            for field in metadata_fields:
                if field in existing_metadata:
                    row = metadata_fields.index(field) + 3
                    ws[f'B{row}'] = existing_metadata[field]

            ws['J3'] = existing_metadata.get('Expediente Físico', '')
            ws['J5'] = existing_metadata.get('No. Carpetas', '')
            ws['J6'] = existing_metadata.get('No. Carpetas Digitalizadas', '')

        # Insertar filas necesarias
        ws.insert_rows(12, amount=len(df))

        # Llenar la plantilla con los datos del DataFrame
        for r, row in enumerate(df.itertuples(index=False), start=12):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)

        # Guardar el archivo
        output_path = os.path.join(folder_path, "000IndiceElectronicoC01.xlsm")
        wb.save(output_path)

        return output_path

def update_metadata(df, metadata):
    """
    Actualiza los metadatos del expediente en el DataFrame.
    """
    metadata_row = pd.DataFrame([metadata])
    return pd.concat([metadata_row, df]).reset_index(drop=True)