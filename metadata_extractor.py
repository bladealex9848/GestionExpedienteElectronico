import os
import PyPDF2
import docx
import openpyxl
from PIL import Image
import magic
from datetime import datetime

def get_file_metadata(file_path):
    """
    Obtiene los metadatos de un archivo.
    """
    stat = os.stat(file_path)
    file_type = get_file_type(file_path)
    metadata = {
        'file_type': file_type,
        'size': format_file_size(stat.st_size),
        'creation_date': format_date(stat.st_ctime),
        'modification_date': format_date(stat.st_mtime),
        'extension': os.path.splitext(file_path)[1].lower()
    }

    if file_type.startswith('application/pdf'):
        metadata.update(get_pdf_metadata(file_path))
    elif file_type.startswith('application/vnd.openxmlformats-officedocument.wordprocessingml'):
        metadata.update(get_word_metadata(file_path))
    elif file_type.startswith('application/vnd.openxmlformats-officedocument.spreadsheetml'):
        metadata.update(get_excel_metadata(file_path))
    elif file_type.startswith('image'):
        metadata.update(get_image_metadata(file_path))

    return metadata

def get_pdf_pages(file_path):
    """
    Obtiene el número de páginas de un archivo PDF.
    """
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            return len(pdf_reader.pages)
    except:
        return 1  # Si no es un PDF o hay un error, asumimos 1 página

def get_pdf_metadata(file_path):
    """
    Obtiene metadatos específicos de archivos PDF.
    """
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            info = pdf_reader.metadata
            return {
                'pages': len(pdf_reader.pages),
                'author': info.author if info.author else 'No disponible',
                'creator': info.creator if info.creator else 'No disponible',
                'producer': info.producer if info.producer else 'No disponible',
                'subject': info.subject if info.subject else 'No disponible',
                'title': info.title if info.title else 'No disponible'
            }
    except:
        return {'pages': 1, 'error': 'No se pudo extraer metadatos del PDF'}

def get_word_metadata(file_path):
    """
    Obtiene metadatos específicos de archivos Word.
    """
    try:
        doc = docx.Document(file_path)
        core_properties = doc.core_properties
        return {
            'pages': len(doc.sections),
            'author': core_properties.author if core_properties.author else 'No disponible',
            'created': format_date(core_properties.created) if core_properties.created else 'No disponible',
            'modified': format_date(core_properties.modified) if core_properties.modified else 'No disponible',
            'title': core_properties.title if core_properties.title else 'No disponible'
        }
    except:
        return {'pages': 1, 'error': 'No se pudo extraer metadatos del documento Word'}

def get_excel_metadata(file_path):
    """
    Obtiene metadatos específicos de archivos Excel.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        return {
            'sheets': len(workbook.sheetnames),
            'sheet_names': ', '.join(workbook.sheetnames)
        }
    except:
        return {'sheets': 1, 'error': 'No se pudo extraer metadatos del archivo Excel'}

def get_image_metadata(file_path):
    """
    Obtiene metadatos específicos de archivos de imagen.
    """
    try:
        with Image.open(file_path) as img:
            return {
                'format': img.format,
                'mode': img.mode,
                'size': f"{img.width}x{img.height}"
            }
    except:
        return {'error': 'No se pudo extraer metadatos de la imagen'}

def get_file_type(file_path):
    """
    Obtiene el tipo MIME del archivo.
    """
    mime = magic.Magic(mime=True)
    return mime.from_file(file_path)

def format_file_size(size):
    """
    Formatea el tamaño del archivo en una unidad legible.
    """
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if size < 1024.0:
            return f"{size:.1f} {unit}"
        size /= 1024.0

def format_date(timestamp):
    """
    Formatea una marca de tiempo en una cadena de fecha legible.
    """
    return datetime.fromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')

def is_document_digitalized(file_path):
    """
    Determina si un documento es digitalizado o nativo electrónico.
    """
    file_type = get_file_type(file_path)
    if file_type.startswith('application/pdf'):
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                if '/Producer' in pdf_reader.metadata:
                    return 'scan' in pdf_reader.metadata['/Producer'].lower()
        except:
            pass
    return False  # Si no es un PDF o no se puede determinar, asumimos que es nativo electrónico