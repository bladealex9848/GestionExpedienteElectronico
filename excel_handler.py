import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import io
import xlwings as xw
import os
import shutil

def save_excel_file(df, file_path, use_template=False, metadata=None):
    """
    Guarda el DataFrame en un archivo Excel, ya sea usando una plantilla o creando uno nuevo.
    
    :param df: DataFrame con los datos del índice
    :param file_path: Ruta donde se guardará el archivo Excel
    :param use_template: Boolean indicando si se debe usar la plantilla
    :param metadata: Diccionario con los metadatos del expediente
    """
    if use_template:
        fill_template_xlwings(df, file_path, metadata)
    else:
        wb = create_new_excel(df, metadata)
        wb.save(file_path)

def create_new_excel(df, metadata=None):
    """
    Crea un nuevo archivo Excel con el formato requerido.
    
    :param df: DataFrame con los datos del índice
    :param metadata: Diccionario con los metadatos del expediente
    :return: Workbook de openpyxl
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Índice Electrónico"

    # Agregar logo
    img = Image('assets/logo.png')
    
    # Ajustar tamaño de la imagen W. 1.9" x H. 0.5"
    img.width = 1.9 * 72
    img.height = 0.5 * 72   

    # Insertar imagen en la celda A1
    ws.add_image(img, 'A1')

    # Título
    ws['B2'] = "ÍNDICE ELECTRÓNICO DEL EXPEDIENTE JUDICIAL"
    ws['B2'].font = Font(bold=True, size=14)
    ws.merge_cells('B2:K2')

    # Metadatos del expediente
    metadata_fields = [
        "Ciudad", "Despacho Judicial", "Serie o Subserie Documental",
        "No. Radicación del Proceso", "Partes Procesales (Parte A)",
        "Partes Procesales (Parte B)", "Terceros Intervinientes", "Cuaderno"
    ]
    for i, field in enumerate(metadata_fields, start=3):
        ws.cell(row=i, column=1, value=field)
        ws.cell(row=i, column=1).font = Font(bold=True)
        ws.merge_cells(f'B{i}:K{i}')
        if metadata and field in metadata:
            ws.cell(row=i, column=2, value=metadata[field])

    # Expediente físico
    ws['L3'] = "EXPEDIENTE FÍSICO"
    ws['L3'].font = Font(bold=True)
    ws.merge_cells('L3:N3')
    ws['L4'] = "El expediente judicial posee documentos físicos:"
    ws['L5'] = "No. de carpetas (cuadernos), legajos o tomos:"
    ws['L6'] = "No. de carpetas (cuadernos), legajos o tomos digitalizados:"
    if metadata:
        ws['J3'] = metadata.get('Expediente Físico', '')
        ws['J5'] = metadata.get('No. Carpetas', '')
        ws['J6'] = metadata.get('No. Carpetas Digitalizadas', '')

    # Encabezados de la tabla
    headers = [
        "Nombre Documento", "Fecha Creación Documento", "Fecha Incorporación Expediente",
        "Orden Documento", "Número Páginas", "Página Inicio", "Página Fin", 
        "Formato", "Tamaño", "Origen", "Observaciones"
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Agregar datos
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    # Aplicar estilos
    for row in ws.iter_rows(min_row=12, max_row=ws.max_row, min_col=1, max_col=11):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    return wb

def fill_template_xlwings(df, file_path, metadata=None):
    """
    Llena la plantilla Excel con los datos del DataFrame usando xlwings.
    
    :param df: DataFrame con los datos del índice
    :param file_path: Ruta donde se guardará el archivo Excel
    :param metadata: Diccionario con los metadatos del expediente
    """
    template_path = os.path.join('assets', '000IndiceElectronicoC0.xlsm')
    shutil.copy(template_path, file_path)
    
    with xw.App(visible=False) as app:
        wb = xw.Book(file_path)
        ws = wb.sheets[0]
        
        # Llenar metadatos del expediente
        metadata_fields = [
            "Ciudad", "Despacho Judicial", "Serie o Subserie Documental",
            "No. Radicación del Proceso", "Partes Procesales (Parte A)",
            "Partes Procesales (Parte B)", "Terceros Intervinientes", "Cuaderno"
        ]
        if metadata:
            for i, field in enumerate(metadata_fields, start=3):
                if field in metadata:
                    ws.range(f'B{i}').value = metadata[field]
            
            ws.range('J3').value = metadata.get('Expediente Físico', '')
            ws.range('J5').value = metadata.get('No. Carpetas', '')
            ws.range('J6').value = metadata.get('No. Carpetas Digitalizadas', '')
        
        # Llenar datos del índice
        ws.range('A12').options(index=False, headers=False).value = df
        
        wb.save()
        wb.close()

def dataframe_to_excel(df):
    """
    Convierte un DataFrame a un archivo Excel en memoria.
    
    :param df: DataFrame a convertir
    :return: Bytes del archivo Excel
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Índice Electrónico')
    return output.getvalue()