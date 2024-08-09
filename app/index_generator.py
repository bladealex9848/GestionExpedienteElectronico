import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side
from fpdf import FPDF
import io
from datetime import datetime

def generate_index(folder_path, metadata):
    """
    Genera el índice electrónico del expediente.
    """
    from expediente_processor import ExpedienteProcessor
    
    processor = ExpedienteProcessor(folder_path)
    return processor.process_with_metadata(metadata)

def generate_excel(df):
    """
    Genera el archivo Excel del índice electrónico.
    """
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Índice Electrónico"

    # Estilos
    header_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))

    # Escribir metadatos del expediente
    for i, (key, value) in enumerate(df.iloc[0].items(), start=1):
        sheet.cell(row=i, column=1, value=key)
        sheet.cell(row=i, column=2, value=value)
        sheet.cell(row=i, column=1).alignment = header_style
        sheet.cell(row=i, column=2).alignment = Alignment(wrap_text=True)

    # Agregar una fila en blanco
    sheet.append([])

    # Escribir datos del índice
    for r in dataframe_to_rows(df.iloc[1:], index=False, header=True):
        sheet.append(r)

    # Aplicar estilos
    for row in sheet.iter_rows(min_row=len(df.iloc[0])+2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.border = border_style

    # Ajustar ancho de columnas
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

    workbook.save(output)
    output.seek(0)  # Asegurarse de que el puntero está al principio
    return output

def generate_pdf(df):
    """
    Genera el archivo PDF del índice electrónico.
    """
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Metadatos del expediente
    pdf.cell(0, 10, "Metadatos del Expediente", 0, 1, 'C')
    pdf.ln(5)
    for key, value in df.iloc[0].items():
        pdf.cell(0, 10, f"{key}: {value}", 0, 1)
    pdf.ln(10)

    # Datos del índice
    pdf.cell(0, 10, "Índice Electrónico", 0, 1, 'C')
    pdf.ln(5)

    # Encabezados de columna
    col_width = pdf.w / 8
    pdf.set_font("Arial", 'B', 10)
    for header in df.columns:
        pdf.cell(col_width, 7, str(header), 1, 0, 'C')
    pdf.ln()

    # Datos
    pdf.set_font("Arial", size=8)
    for i, row in df.iloc[1:].iterrows():
        for item in row:
            pdf.cell(col_width, 6, str(item), 1)
        pdf.ln()

    # Pie de página
    pdf.set_y(-15)
    pdf.set_font('Arial', 'I', 8)
    pdf.cell(0, 10, f'Generado el {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 0, 0, 'C')

    output = io.BytesIO()
    pdf.output(output, 'S')  # Cambiar 'F' a 'S' para escribir en BytesIO
    output.seek(0)  # Asegurarse de que el puntero está al principio
    return output